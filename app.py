import os
import uuid
import pandas as pd
import sqlite3
from flask import Flask, request, render_template, redirect, url_for, flash, send_file
from jinja2 import Environment, FileSystemLoader
from werkzeug.utils import secure_filename
import smtplib
from email.message import EmailMessage
from datetime import datetime
import pytz

app = Flask(__name__)
app.secret_key = "your_secret_key"

UPLOAD_FOLDER = "uploads"
HTML_FOLDER = "static/confirm_pages"
DB_PATH = "submissions.db"
ALLOWED_EXTENSIONS = {"xlsx"}

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(HTML_FOLDER, exist_ok=True)

# 建立 SQLite 資料庫
def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS submissions (
        id TEXT PRIMARY KEY,
        name TEXT,
        email TEXT,
        title TEXT,
        fee INTEGER,
        bank TEXT,
        account TEXT,
        account_name TEXT,
        submitted_at TEXT
    )""")
    conn.commit()
    conn.close()

init_db()

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route("/", methods=["GET", "POST"])
def upload_file():
    if request.method == "POST":
        file = request.files["file"]
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            path = os.path.join(UPLOAD_FOLDER, filename)
            file.save(path)
            generate_confirm_pages(path)
            flash("成功上傳並產生確認頁！")
            return redirect(url_for("upload_file"))
        else:
            flash("請上傳 .xlsx 檔案")
            return redirect(url_for("upload_file"))
    return render_template("upload.html")

def generate_confirm_pages(filepath):
    df = pd.read_excel(filepath)

    # 若沒有 id 欄位，自動新增
    if "id" not in df.columns:
        df.insert(0, "id", [str(uuid.uuid4())[:8] for _ in range(len(df))])
        df.to_excel(filepath, index=False)

    env = Environment(loader=FileSystemLoader("templates"))
    template = env.get_template("confirm_template.html")

    for _, row in df.iterrows():
        unique_id = row["id"]
        filename = f"{unique_id}.html"
        html = template.render(
            id=unique_id,
            name=row["Author"],
            email=row["E-mail"],
            title=row["Title"],
            fee=row["Fee"]
        )
        with open(os.path.join(HTML_FOLDER, filename), "w", encoding="utf-8") as f:
            f.write(html)
        send_email(row["Author"], row["E-mail"], unique_id)

def send_email(name, recipient, page_id):
    msg = EmailMessage()
    msg["Subject"] = "【幻華創造】稿費資訊確認通知信"
    msg["From"] = os.getenv("EMAIL_ADDRESS")
    msg["To"] = recipient
    link = f"{os.getenv('BASE_URL')}/static/confirm_pages/{page_id}.html"
    msg.set_content(f"您好 {name}，\n\n請點選以下連結，確認您的稿費資訊是否正確，並填寫帳戶資料：\n{link}")

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
        smtp.login(os.getenv("EMAIL_ADDRESS"), os.getenv("EMAIL_PASSWORD"))
        smtp.send_message(msg)

@app.route("/submit", methods=["POST"])
def submit():
    data = request.form
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    # 查詢該 ID 是否已經存在
    c.execute("SELECT * FROM submissions WHERE id = ?", (data["id"],))
    existing = c.fetchone()

    if existing:
        conn.close()
        return "⚠️ 您已經填寫過表單了，無需重複提交。"

    submitted_at = datetime.now(pytz.timezone("Asia/Taipei")).strftime("%Y/%m/%d %H:%M")
    account_name = data.get("account_name")

    c.execute("INSERT INTO submissions (id, name, email, title, fee, bank, account, account_name, submitted_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)", (
        data["id"], data["name"], data["email"], data["title"],
        data["fee"], data["bank"], data["account"], account_name, submitted_at
    ))
    conn.commit()
    conn.close()
    return "✅ 表單已成功送出，感謝您！"

@app.route("/export")
def export_page():
    return render_template("export.html")

@app.route("/download-export")
def download_export():
    conn = sqlite3.connect(DB_PATH)
    submitted_df = pd.read_sql_query("SELECT * FROM submissions", conn)
    submitted_ids = submitted_df["id"].tolist()
    conn.close()

    # 找出最新的 authors.xlsx
    files = sorted([f for f in os.listdir(UPLOAD_FOLDER) if f.endswith(".xlsx")])
    if not files:
        return "❗ 無可比對的 authors.xlsx 檔案。請先上傳。"
    latest_file = os.path.join(UPLOAD_FOLDER, files[-1])
    df_authors = pd.read_excel(latest_file)

    pending_df = df_authors[~df_authors["id"].isin(submitted_ids)]

    # 在兩個表格中加上「狀態」欄
    submitted_df.insert(0, "狀態", "✅ 已回覆")
    pending_df.insert(0, "狀態", "⏳ 未回覆")

    # 重新命名欄位順序（若有不同）
    # 補上缺的欄位再重新排序
    for col in submitted_df.columns:
        if col not in pending_df.columns:
            pending_df[col] = ""
    pending_df = pending_df[submitted_df.columns]

    combined_df = pd.concat([submitted_df, pending_df], ignore_index=True)

    export_path = "submissions_combined.xlsx"
    export_to_excel_pretty(combined_df, export_path)
    return send_file(export_path, as_attachment=True)

def export_to_excel_pretty(df, export_path="submissions_export.xlsx"):
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font
    with pd.ExcelWriter(export_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="回覆紀錄")
        ws = writer.sheets["回覆紀錄"]
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max_length + 2

@app.route("/pending")
def pending_list():
    conn = sqlite3.connect(DB_PATH)
    submitted_ids = pd.read_sql_query("SELECT id FROM submissions", conn)["id"].tolist()
    conn.close()

    files = sorted([f for f in os.listdir(UPLOAD_FOLDER) if f.endswith(".xlsx")])
    if not files:
        return "❗ 無可比對的 authors.xlsx 檔案。請先上傳。"

    latest_file = os.path.join(UPLOAD_FOLDER, files[-1])
    df_authors = pd.read_excel(latest_file)

    pending_df = df_authors[~df_authors["id"].isin(submitted_ids)]

    export_path = "pending.xlsx"
    export_to_excel_pretty(pending_df, export_path)
    return send_file(export_path, as_attachment=True)

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 10000)))
